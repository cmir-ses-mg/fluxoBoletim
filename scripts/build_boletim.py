# -*- coding: utf-8 -*-
"""
build_boletim.py — Gera o Boletim de Acompanhamento do Plano Rio Doce
Lê a Base Consolidada (xlsx) e produz um boletim.html autocontido, pronto
para impressão em A4 (Ctrl+P) ou publicação em página estática.
Uso: python build_boletim.py [caminho_do_xlsx] [saida.html]
"""
import sys, datetime, unicodedata
from collections import defaultdict, Counter
import openpyxl

import glob, os

def achar_planilha(alvo):
    """Aceita um arquivo .xlsx OU uma pasta: nesse caso pega o .xlsx mais recente,
    ignorando temporários do Excel (~$...). Assim o nome do arquivo pode mudar
    livremente — inclusive com prefixos como [CONSOLIDADO]."""
    if os.path.isfile(alvo):
        return alvo
    if os.path.isdir(alvo):
        cands = [f for f in glob.glob(os.path.join(alvo, "*.xlsx"))
                 if not os.path.basename(f).startswith("~$")]
        if not cands:
            sys.exit(f"ERRO: nenhum .xlsx encontrado em '{alvo}'. "
                     "Suba a planilha consolidada nessa pasta e rode de novo.")
        mais_novo = max(cands, key=os.path.getmtime)
        if len(cands) > 1:
            print(f"[aviso] {len(cands)} planilhas na pasta — usando a mais recente: {os.path.basename(mais_novo)}")
        return mais_novo
    sys.exit(f"ERRO: caminho não encontrado: '{alvo}'")

XLSX = achar_planilha(sys.argv[1] if len(sys.argv) > 1 else "dados")
OUT  = sys.argv[2] if len(sys.argv) > 2 else "boletim.html"
print(f"Planilha de origem: {os.path.basename(XLSX)}")

def brl(v, dec=2):
    s = f"{v:,.{dec}f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"R$ {s}"

def w(x):
    """Largura em % para CSS, limitada a 0–100."""
    return f"{max(0, min(100, x)):.1f}"


def minibar(p):
    return f"<div class='mini'><div class='mini-fill' style='width:{w(p)}%'></div></div>"


def pct(x, dec=1):
    return f"{x:.{dec}f}".replace(".", ",") + "%"

# ── LEITURA ──────────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(XLSX, data_only=True)
for aba in ("Base Consolidada", "Atualizações do Plano"):
    if aba not in wb.sheetnames:
        sys.exit(f"ERRO: a planilha não tem a aba '{aba}'. Abas encontradas: {', '.join(wb.sheetnames)}")
ws = wb["Base Consolidada"]
H = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]


def _chave(s):
    """Normaliza cabeçalho: sem acento, sem pontuação, minúsculo, espaços colapsados.
    'Status  pré-repasse ' e 'Status pre repasse' viram a mesma chave."""
    s = unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode()
    s = "".join(ch if ch.isalnum() or ch.isspace() else " " for ch in s)
    return " ".join(s.lower().split())


_HMAP = {}
for _i, _h in enumerate(H, start=1):
    if _h is not None and str(_h).strip():
        _HMAP.setdefault(_chave(_h), _i)

# nome usado no script -> nomes aceitos na planilha (o primeiro encontrado vence)
ALIASES = {
    "Valor": ["Valor", "Valor Previsto", "Valor - DMPS", "Valor - Fonte 80"],
    "Valor Pago": ["Valor Pago", "valor pago"],
    "Ação": ["Ação", "Ação do Plano"],
    "N° ação": ["N° ação", "Nº ação", "No ação", "N ação"],
    "Beneficiário": ["Beneficiário"],
    "URS": ["URS", "Unidade Regional de Saúde"],
    "Tipo de Aplicação": ["Tipo de Aplicação"],
    "Instrumento de execução do recurso": ["Instrumento de execução do recurso", "Instrumento"],
    "Status do pagamento": ["Status do pagamento"],
    "Status do Repasse": ["Status do Repasse"],
    "Status pré-repasse": ["Status pré-repasse", "Status Obras (pré-repasse)",
                           "Status de Obras (pré-repasse)", "Status de obras",
                           "Status de custeio (pré-repasse)"],
    "Status de equipamentos (pré-repasse)": ["Status de equipamentos (pré-repasse)"],
    "Status de celebração de convênio (pré-repasse)": ["Status de celebração de convênio (pré-repasse)"],
    "SEI-OBRAS": ["SEI-OBRAS", "SEI OBRAS"],
    "Beneficiário Final": ["Beneficiário Final"],
    "Data da atualização": ["Data da atualização", "Data de atualização", "Data da atualizacao"],
    "MICRO": ["MICRO", "Microrregião", "Microrregiao", "Micro"],
    "Status de celebração de contrato (pré-repasse)": ["Status de celebração de contrato (pré-repasse)"],
    "Status dos Termos de Adesão": ["Status dos Termos de Adesão", "Status dos Termos de Adesao"],
    "Status de Resolução - alteras": ["Status de Resolução - alteras", "Status de Resolução"],
}

IX = {}
for _canon, _ops in ALIASES.items():
    for _op in _ops:
        if _chave(_op) in _HMAP:
            IX[_canon] = _HMAP[_chave(_op)]
            break

OBRIGATORIAS = ["Valor", "Valor Pago", "Beneficiário", "URS", "N° ação", "Tipo de Aplicação",
                "Instrumento de execução do recurso", "Status do pagamento", "Status do Repasse",
                "Status pré-repasse", "Status de equipamentos (pré-repasse)",
                "Status de celebração de convênio (pré-repasse)"]
_faltando = [c for c in OBRIGATORIAS if c not in IX]
if _faltando:
    sys.exit("ERRO: não encontrei estas colunas na aba 'Base Consolidada':\n  - "
             + "\n  - ".join(_faltando)
             + "\n\nColunas existentes na planilha:\n  "
             + "\n  ".join(str(h) for h in H if h)
             + "\n\nRenomeie a coluna na planilha ou acrescente o nome em ALIASES, no script.")

for _canon in ALIASES:
    if _canon in IX and str(H[IX[_canon] - 1]).strip() != _canon:
        print(f"[info] coluna '{H[IX[_canon]-1]}' reconhecida como '{_canon}'")


def g(r, n):
    return str(ws.cell(r, IX[n]).value or "").strip() if n in IX else ""


def gv(r, n):
    return ws.cell(r, IX[n]).value if n in IX else None

rows = range(2, ws.max_row + 1)
OBRAS_ACOES = (1, 2, 6)

tot = pago = emp = 0.0
by_instr = defaultdict(lambda: [0.0, 0.0])
by_urs   = defaultdict(lambda: [0.0, 0.0])
benef    = defaultdict(lambda: [0.0, 0.0])
pg_int = pg_par = pg_emp = pg_sem = 0
obras_status = defaultdict(Counter); obras_val_travado = 0.0
equip_ok = equip_at = equip_na = 0; equip_at_quem = []
conv_grupos = defaultdict(lambda: [0, 0.0, []])
parciais = []
n_ubs = n_caps = 0

for r in rows:
    v  = gv(r, "Valor") or 0
    vp = gv(r, "Valor Pago")
    sp = g(r, "Status do pagamento")
    pg = vp if vp is not None else (v if sp == "Liquidado/Pago" else 0)
    rep, instr = g(r, "Status do Repasse"), g(r, "Instrumento de execução do recurso")
    ben, urs = g(r, "Beneficiário"), g(r, "URS")
    na = gv(r, "N° ação")

    tot += v; pago += pg
    if rep == "Empenhado": emp += max(0, v - pg)
    by_instr[instr][0] += v; by_instr[instr][1] += pg
    by_urs[urs][0] += v;     by_urs[urs][1] += pg
    benef[ben][0] += v;      benef[ben][1] += pg

    if sp == "Liquidado/Pago": pg_int += 1
    elif sp == "Liquidado/Pago PARCIAL":
        pg_par += 1
        parciais.append((ben, g(r, "Tipo de Aplicação"), v, pg))
    elif sp == "Empenhado": pg_emp += 1
    else: pg_sem += 1

    if instr == "Resolução - Investimento" and na in OBRAS_ACOES:
        st = g(r, "Status pré-repasse") or "Não iniciado"
        obras_status[st][ben] += 1
        tipo = g(r, "Tipo de Aplicação")
        if "UBS" in tipo: n_ubs += 1
        if "CAPS" in tipo: n_caps += 1
        if st in ("Aguardando envio dos documentos", "[CAPS] Aguardando PAR da RAPS", "Pleito em reavaliação"):
            obras_val_travado += v
    if instr == "Resolução - Investimento" and na not in OBRAS_ACOES:
        se = g(r, "Status de equipamentos (pré-repasse)")
        if se == "Lista validada": equip_ok += 1
        elif se == "Não se aplica" or not se or se == "-": equip_na += 1
        else:
            equip_at += 1; equip_at_quem.append(f"{ben} ({se})")
    if instr == "Convênio":
        sc = g(r, "Status de celebração de convênio (pré-repasse)") or "Pendente"
        conv_grupos[sc][0] += 1; conv_grupos[sc][1] += v; conv_grupos[sc][2].append(ben)

# ── por instrumento: previsto, pago, empenhado e etapas em curso ──
INSTR_ORDEM = ["Resolução - Investimento", "Convênio", "Resolução - Custeio", "Contrato"]
INSTR_LABEL = {"Resolução - Investimento": "Resolução — Investimento",
               "Resolução - Custeio": "Resolução — Custeio",
               "Convênio": "Convênio", "Contrato": "Contrato"}
INSTR_FINANCIA = {
    "Resolução - Investimento": "{n_ubs} UBS e {n_caps} CAPS; equipamentos hospitalares, de APS e de consórcios",
    "Convênio": "obras de maior porte — hospitais, CEAE, polos de fisioterapia e SAE",
    "Resolução - Custeio": "VIGIAGUA, VIGIAR e VIGIDESASTRES; custeio hospitalar e de UBS",
    "Contrato": "vigilância de agravos do desastre e monitoramento técnico de obras",
}

def situacao_linha(r):
    """Situação do processo na linha — a etapa em que a ação está.
    Não usa o status de pagamento: procura o status próprio do instrumento
    (obras, equipamentos, celebração) e, se estiver vazio, cai para o termo
    de adesão ou para a resolução."""
    instr_ = g(r, "Instrumento de execução do recurso")
    na_ = gv(r, "N° ação")
    candidatos = []
    if instr_ == "Resolução - Investimento":
        if na_ in OBRAS_ACOES:
            candidatos = [("Obras", "Status pré-repasse")]
        else:
            candidatos = [("Equipamentos", "Status de equipamentos (pré-repasse)")]
    elif instr_ == "Convênio":
        candidatos = [("Celebração", "Status de celebração de convênio (pré-repasse)")]
    elif instr_ == "Contrato":
        candidatos = [("Celebração", "Status de celebração de contrato (pré-repasse)")]
    elif instr_ == "Resolução - Custeio":
        candidatos = [("Pré-repasse", "Status pré-repasse"),
                      ("Resolução", "Status de Resolução - alteras")]
    candidatos += [("Termo de adesão", "Status dos Termos de Adesão"),
                   ("Repasse", "Status do Repasse")]
    for etapa_, col in candidatos:
        val = g(r, col)
        if val and val not in ("-", "–", "Não se aplica"):
            return (val, etapa_)
    return ("—", "")


instr_dados = defaultdict(lambda: {"prev": 0.0, "pago": 0.0, "emp": 0.0, "n": 0})
instr_etapas = defaultdict(lambda: defaultdict(lambda: [0, 0.0, Counter()]))
CRIT_OBRAS = ("Aguardando envio dos documentos", "[CAPS] Aguardando PAR da RAPS", "Pleito em reavaliação")
CRIT_CONV = ("Pleito em reavaliação", "Comunicado DCR")

por_micro = defaultdict(lambda: [0.0, 0.0])
por_benef_mapa = defaultdict(lambda: [0.0, 0.0])
atualizacoes = []
hoje_d = datetime.date.today()
JANELA_DIAS = 7   # dias corridos, contando o dia de hoje

for r in rows:
    v = gv(r, "Valor") or 0
    vp = gv(r, "Valor Pago")
    sp = g(r, "Status do pagamento")
    pg = vp if vp is not None else (v if sp == "Liquidado/Pago" else 0)
    rep = g(r, "Status do Repasse")
    instr = g(r, "Instrumento de execução do recurso")
    na = gv(r, "N° ação")

    d = instr_dados[instr]
    d["prev"] += v
    d["pago"] += pg
    d["n"] += 1
    if rep == "Empenhado":
        d["emp"] += max(0, v - pg)

    # etapas em curso (o que estava em "Em tramitação regular")
    etapa = ""
    if instr == "Resolução - Investimento":
        if na in OBRAS_ACOES:
            st = g(r, "Status pré-repasse")
            if st and st not in CRIT_OBRAS and st != "Análise técnica liberada":
                etapa = f"Obras — {st}"
        else:
            st = g(r, "Status de equipamentos (pré-repasse)")
            if st and st not in ("Lista validada", "Não se aplica", "-", "Em ajuste", "Aguardando envio da lista"):
                etapa = f"Equipamentos — {st}"
    elif instr == "Convênio":
        st = g(r, "Status de celebração de convênio (pré-repasse)")
        if st and st not in CRIT_CONV and st != "Assinatura e publicação do convênio":
            etapa = st
    elif instr == "Contrato":
        st = g(r, "Status de celebração de contrato (pré-repasse)")
        if st and st not in ("-", "Assinado"):
            etapa = st
    if etapa:
        e = instr_etapas[instr][etapa]
        e[0] += 1
        e[1] += v
        e[2][g(r, "Beneficiário")] += 1

    # execução por município (alimenta o mapa)
    ben = g(r, "Beneficiário")
    por_benef_mapa[ben][0] += v
    por_benef_mapa[ben][1] += pg

    # execução por microrregião
    micro = g(r, "MICRO") or "Não informada"
    por_micro[micro][0] += v
    por_micro[micro][1] += pg

    # atualizações recentes
    dt = gv(r, "Data da atualização")
    if dt is not None:
        dd = dt.date() if hasattr(dt, "date") else dt
        try:
            # janela de 7 dias corridos INCLUINDO hoje (hoje-6 ... hoje);
            # datas futuras (erro de digitação) ficam de fora
            if 0 <= (hoje_d - dd).days <= JANELA_DIAS - 1:
                atualizacoes.append((dd, g(r, "Beneficiário"), g(r, "Beneficiário Final"),
                                     g(r, "Tipo de Aplicação")) + situacao_linha(r))
        except TypeError:
            pass

atualizacoes.sort(key=lambda x: (-x[0].toordinal(), x[1]))

demais = max(0, tot - pago - emp)
comprometido = pago + emp
p_pago, p_emp, p_dem = 100*pago/tot, 100*emp/tot, 100*demais/tot
p_compr = 100*comprometido/tot
n_benef = len(benef)
n_linhas = ws.max_row - 1
n_obras = n_ubs + n_caps

# obras: agrupamento crítico / análise
ob_crit = {k: v for k, v in obras_status.items() if k in ("Aguardando envio dos documentos", "[CAPS] Aguardando PAR da RAPS", "Pleito em reavaliação", "Não iniciado")}
ob_anda = {k: v for k, v in obras_status.items() if k not in ob_crit}
ob_crit_n = sum(sum(c.values()) for c in ob_crit.values())
ob_anda_n = sum(sum(c.values()) for c in ob_anda.values())

def lista_munic(counter):
    return ", ".join(f"{m} ({q})" if q > 1 else m for m, q in sorted(counter.items()))

def lista_conv(nomes):
    c = Counter(nomes)
    return ", ".join(f"{m} ({q})" if q > 1 else m for m, q in sorted(c.items()))

# convênios: críticos vs tramitação
CONV_CRIT = ("Pleito em reavaliação", "Comunicado DCR", "Pendente")
conv_crit_n  = sum(v[0] for k, v in conv_grupos.items() if k in CONV_CRIT)
conv_crit_val = sum(v[1] for k, v in conv_grupos.items() if k in CONV_CRIT)
conv_tram_n  = sum(v[0] for k, v in conv_grupos.items() if k not in CONV_CRIT)
conv_tram_val = sum(v[1] for k, v in conv_grupos.items() if k not in CONV_CRIT)
conv_total_n = conv_crit_n + conv_tram_n
conv_total_val = by_instr["Convênio"][0]

# alterações
wa = wb["Atualizações do Plano"]
HA = [wa.cell(1, c).value for c in range(1, wa.max_column + 1)]
def ga(r, n): return str(wa.cell(r, HA.index(n) + 1).value or "").strip()
alt_fin, alt_and, alt_and_lista = 0, 0, []
for r in range(2, wa.max_row + 1):
    if not wa.cell(r, 1).value: continue
    st = ga(r, "Status da alteração")
    if st == "Finalizado": alt_fin += 1
    elif st == "Em andamento":
        alt_and += 1
        alt_and_lista.append(f"{ga(r,'Beneficiário')} ({ga(r,'Categoria de alteração de pleito').lower() or 'objeto'})")

# instrumentos em ordem de valor
INSTR_LABEL = {"Resolução - Investimento": "Resolução — Investimento",
               "Resolução - Custeio": "Resolução — Custeio",
               "Convênio": "Convênio", "Contrato": "Contrato"}
instr_rows = sorted(by_instr.items(), key=lambda x: -x[1][0])

INSTR_NOTA = {
    "Resolução - Investimento": f"{n_ubs} UBS e {n_caps} CAPS; equipamentos hospitalares, de APS e de consórcios.",
    "Resolução - Custeio": "VIGIAGUA, VIGIAR e VIGIDESASTRES; custeio hospitalar e de UBS.",
    "Convênio": "Obras de maior porte — hospitais, CEAE, polos de fisioterapia e SAE.",
    "Contrato": "Vigilância de agravos do desastre e monitoramento técnico de obras.",
}

urs_rows = sorted(((k, v) for k, v in by_urs.items()), key=lambda x: -x[1][0])
benef_100 = sorted([b for b, (v, p) in benef.items() if v > 0 and p >= v - 0.01])
n_benef_pagos = sum(1 for b, (v, p) in benef.items() if p > 0)

# ── cards de instrumento (barra + o que financia + etapas em curso) ──
cards_instr = ""
for k in INSTR_ORDEM:
    if k not in instr_dados:
        continue
    d = instr_dados[k]
    prev, pgo, empn = d["prev"], d["pago"], d["emp"]
    x = 100 * pgo / prev if prev else 0
    xe = 100 * empn / prev if prev else 0
    etapas_txt = ""
    for et, (n, val, munis) in sorted(instr_etapas[k].items(), key=lambda x: -x[1][1]):
        quem = ", ".join(f"{m} ({q})" if q > 1 else m for m, q in sorted(munis.items()))
        etapas_txt += (f"<div class='et'><span class='et-n'>{n}</span>"
                       f"<span class='et-d'>{et}{' · ' + brl(val, 0) if val else ''}"
                       f"<span class='et-m'>{quem}</span></span></div>")
    bloco_etapas = (f"<div class='et-wrap'><div class='et-lbl'>Etapas em curso</div>{etapas_txt}</div>"
                    if etapas_txt else "")
    cards_instr += f"""<div class='ci'>
      <div class='ci-top'>
        <div class='ci-nome'>{INSTR_LABEL.get(k, k)}<span class='ci-n'>{d['n']} ações</span></div>
        <div class='ci-prev'>{brl(prev, 0)}</div>
      </div>
      <div class='ci-bar'>
        <div class='ci-pago' style='width:{w(x)}%'></div>
        <div class='ci-emp' style='width:{w(xe)}%'></div>
      </div>
      <div class='ci-leg'>
        <span><i class='dot dp'></i><b>{brl(pgo, 0)}</b> pago ({pct(x, 0)})</span>
        <span><i class='dot de'></i><b>{brl(empn, 0)}</b> em tramitação</span>
        <span class='ci-fin'>Financia {INSTR_FINANCIA.get(k, '').format(n_ubs=n_ubs, n_caps=n_caps)}</span>
      </div>
      {bloco_etapas}
    </div>"""

# ── atualizações da semana ──
linhas_atu = "".join(
    f"<tr><td class='t-nome'>{ben}</td><td>{bf if bf and bf != '-' else '&mdash;'}</td>"
    f"<td>{tipo}</td>"
    f"<td>{st}{f'<span class=at-et>{etapa_}</span>' if etapa_ else ''}</td>"
    f"<td class='t-num'>{dd.strftime('%d/%m/%Y')}</td></tr>"
    for dd, ben, bf, tipo, st, etapa_ in atualizacoes)

# ── execução por microrregião ──
linhas_micro = ""
for m, (prev, pgo) in sorted(por_micro.items(), key=lambda x: -x[1][0]):
    x = 100 * pgo / prev if prev else 0
    linhas_micro += (f"<tr><td class='t-nome'>{m}</td><td class='t-num'>{brl(prev)}</td>"
                     f"<td class='t-num'>{brl(pgo)}</td>"
                     f"<td class='t-exec'><span class='t-pct'>{pct(x)}</span>{minibar(x)}</td></tr>")

# ── mapa coroplético: cada município pintado pelo % executado ────────────────
# Malha oficial do IBGE (recortada só para os municípios do plano).
MALHA = sys.argv[3] if len(sys.argv) > 3 else os.path.join("assets", "malha_rio_doce.json")

FAIXAS = [(0.001, "#E7EDE8", "Sem pagamento"),
          (25, "#BFDCCB", "até 25%"),
          (50, "#7FBB9D", "25% a 50%"),
          (75, "#3F9A70", "50% a 75%"),
          (100.01, "#166A47", "acima de 75%")]


def cor_faixa(p):
    for lim, cor, _ in FAIXAS:
        if p < lim:
            return cor
    return FAIXAS[-1][1]


def gerar_mapa(caminho, execucao):
    """SVG com os municípios do plano, pintados pelo percentual executado."""
    if not os.path.isfile(caminho):
        print(f"[aviso] malha não encontrada em '{caminho}' — seção regional sai sem o mapa")
        return ""
    import json as _json
    with open(caminho, encoding="utf-8") as fm:
        gj = _json.load(fm)

    def aneis(feat):
        g_ = feat["geometry"]
        if g_["type"] == "Polygon":
            return g_["coordinates"]
        return [anel for poly in g_["coordinates"] for anel in poly]

    xs = [x for f in gj["features"] for a in aneis(f) for x, y in a]
    ys = [y for f in gj["features"] for a in aneis(f) for x, y in a]
    x0, x1, y0, y1 = min(xs), max(xs), min(ys), max(ys)
    import math as _math
    kx0 = _math.cos(_math.radians((y0 + y1) / 2))
    L = 1000
    pad = 10
    A = int((L - 2 * pad) * (y1 - y0) / ((x1 - x0) * kx0)) + 2 * pad
    kx = kx0
    esc = min((L - 2 * pad) / ((x1 - x0) * kx), (A - 2 * pad) / (y1 - y0))
    dx = (L - (x1 - x0) * kx * esc) / 2
    dy = (A - (y1 - y0) * esc) / 2

    def proj(x, y):
        return (dx + (x - x0) * kx * esc, A - dy - (y - y0) * esc)

    partes = []
    for f in gj["features"]:
        nome = f["properties"]["nome"]
        prev, pgo = execucao.get(nome, (0, 0))
        p = 100 * pgo / prev if prev else 0
        d = ""
        for anel in aneis(f):
            pts = [proj(x, y) for x, y in anel]
            d += "M" + " L".join(f"{px:.1f},{py:.1f}" for px, py in pts) + "Z"
        partes.append(f"<path d='{d}' fill='{cor_faixa(p)}' stroke='#fff' stroke-width='.7'>"
                      f"<title>{nome}: {pct(p)} executado</title></path>")

    legenda = "".join(
        f"<span class='mp-item'><i style='background:{cor}'></i>{rot}</span>"
        for _, cor, rot in FAIXAS)
    return (f"<div class='mapa'>"
            f"<svg viewBox='0 0 {L} {A}' xmlns='http://www.w3.org/2000/svg' role='img' "
            f"aria-label='Execução financeira por município'>{''.join(partes)}</svg>"
            f"<div class='mp-leg'>{legenda}</div>"
            f"<div class='mapa-cap'>Execução financeira por município — {len(gj['features'])} municípios do plano. "
            f"Consórcios e SES não aparecem no mapa por não terem território próprio.</div></div>")


mapa_html = gerar_mapa(MALHA, {k: tuple(v) for k, v in por_benef_mapa.items()})

hoje = datetime.date.today()
MESES = ["janeiro","fevereiro","março","abril","maio","junho","julho","agosto","setembro","outubro","novembro","dezembro"]
data_ext = f"{hoje.day} de {MESES[hoje.month-1]} de {hoje.year}"
edicao = f"{hoje.month:02d}/{hoje.year}"

# ── FRAGMENTOS HTML ──────────────────────────────────────────────────────────
linhas_instr = "\n".join(
    f"""<tr>
      <td class='t-nome'>{INSTR_LABEL.get(k,k)}</td>
      <td class='t-num'>{brl(v)}</td>
      <td class='t-num'>{brl(p)}</td>
      <td class='t-exec'><span class='t-pct'>{pct(100*p/v if v else 0)}</span>{minibar(100*p/v if v else 0)}</td>
      <td class='t-nota'>{INSTR_NOTA.get(k,'')}</td>
    </tr>""" for k, (v, p) in instr_rows)

linhas_urs = "\n".join(
    f"""<tr><td class='t-nome'>{k}</td><td class='t-num'>{brl(v)}</td>
      <td class='t-num'>{brl(p)}</td>
      <td class='t-exec'><span class='t-pct'>{pct(100*p/v if v else 0)}</span>{minibar(100*p/v if v else 0)}</td></tr>"""
    for k, (v, p) in urs_rows)

linhas_parciais = "\n".join(
    f"""<tr><td class='t-nome'>{b}</td><td>{t}</td>
      <td class='t-num'>{brl(v)}</td><td class='t-num'>{brl(p)}</td>
      <td class='t-num'>{brl(v-p)}</td></tr>""" for b, t, v, p in sorted(parciais, key=lambda x: -(x[2]-x[3])))

def bloco_pend(titulo, qtd, sub, munics, tom):
    return f"""<div class='pend {tom}'>
      <div class='pend-head'><span class='pend-num'>{qtd}</span><div><div class='pend-t'>{titulo}</div><div class='pend-s'>{sub}</div></div></div>
      <div class='pend-m'>{munics}</div></div>"""

pendencias = ""
if "Aguardando envio dos documentos" in ob_crit:
    c = ob_crit["Aguardando envio dos documentos"]
    pendencias += bloco_pend("Unidades aguardando envio de documentos de obra",
        sum(c.values()), f"em {len(c)} municípios",
        lista_munic(c), "rio")
if "[CAPS] Aguardando PAR da RAPS" in ob_crit:
    c = ob_crit["[CAPS] Aguardando PAR da RAPS"]
    pendencias += bloco_pend("CAPS aguardando aprovação do PAR junto à RAPS",
        sum(c.values()), "condição prévia à análise de engenharia", lista_munic(c), "rio")
pleitos_ob = obras_status.get("Pleito em reavaliação", Counter())
if pleitos_ob or conv_crit_n:
    partes = []
    if pleitos_ob:
        partes.append(("Obra: " if sum(pleitos_ob.values()) == 1 else "Obras: ") + lista_munic(pleitos_ob))
    if conv_crit_n:
        noms = sum((v[2] for k, v in conv_grupos.items() if k in CONV_CRIT), [])
        partes.append(("Convênio: " if conv_crit_n == 1 else "Convênios: ") + lista_conv(noms))
    munics = " &nbsp;·&nbsp; ".join(partes)
    pendencias += bloco_pend("Atualizações do Plano",
        sum(pleitos_ob.values()) + conv_crit_n,
        f"objeto/valor em rediscussão — {brl(conv_crit_val, 0)} em convênios aguardando a definição",
        munics, "rio")

tramitacao = ""
conv_tram_itens = sorted(((k, v) for k, v in conv_grupos.items() if k not in CONV_CRIT), key=lambda x: -x[1][1])
for k, (q, v, noms) in conv_tram_itens:
    tramitacao += f"""<tr><td class='t-nome'>{k}</td><td>Convênio</td><td class='t-num'>{q}</td>
      <td class='t-num'>{brl(v,0)}</td><td class='t-nota'>{lista_conv(noms)}</td></tr>"""
if ob_anda:
    for k, c in sorted(ob_anda.items(), key=lambda x: -sum(x[1].values())):
        tramitacao += f"""<tr><td class='t-nome'>{k}</td><td>Investimento — Obras</td><td class='t-num'>{sum(c.values())}</td>
          <td class='t-num'>—</td><td class='t-nota'>{lista_munic(c)}</td></tr>"""
if equip_at:
    tramitacao += f"""<tr><td class='t-nome'>Listas em análise</td><td>Investimento — Equipamentos</td><td class='t-num'>{equip_at}</td>
      <td class='t-num'>—</td><td class='t-nota'>{", ".join(equip_at_quem)}</td></tr>"""

# ── HTML ─────────────────────────────────────────────────────────────────────
html = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Boletim de Acompanhamento — Plano de Ação em Saúde do Rio Doce</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Archivo:wdth,wght@62..125,400..900&family=Source+Serif+4:ital,opsz,wght@0,8..60,400..700;1,8..60,400..700&display=swap" rel="stylesheet">
<style>
  :root {{
    --tinta:#1C2420; --mata:#1E6B47; --mata-escuro:#0F3D2A;
    --rio:#C2551B; --rio-claro:#FBEFE7;
    --ambar:#9A6A00; --ambar-claro:#FBF3DF;
    --verde-claro:#E8F2EC; --painel:#F2F4F0; --fio:#DCE3DB;
    --cinza:#5C6660; --cinza-claro:#8B948E;
  }}
  * {{ box-sizing:border-box; margin:0; }}
  html {{ -webkit-print-color-adjust:exact; print-color-adjust:exact; }}
  body {{ font-family:'Source Serif 4', 'Cambria', Georgia, serif; color:var(--tinta); background:#E6E9E4;
         font-size:11.5pt; line-height:1.62; }}
  .arch {{ font-family:'Archivo', 'Segoe UI', 'Helvetica Neue', Arial, sans-serif; }}

  .pagina {{ background:#fff; width:210mm; min-height:297mm; margin:10mm auto; padding:16mm 17mm 14mm; display:flex; flex-direction:column;
             box-shadow:0 2px 18px rgba(28,36,32,.14); position:relative; }}

  /* ── assinatura: o fio do rio (divisor mata→barro) ── */
  .fio-rio {{ height:2.5px; border:0; background:linear-gradient(90deg, var(--mata) 0%, var(--mata) 55%, var(--rio) 100%); margin:0 0 5mm; }}

  /* ── masthead ── */
  .cabo {{ display:flex; justify-content:space-between; align-items:flex-end; padding-bottom:4mm; }}
  .cabo-org {{ font-family:'Archivo'; font-size:8pt; font-weight:600; letter-spacing:.14em; text-transform:uppercase; color:var(--cinza); }}
  .cabo-num {{ font-family:'Archivo'; font-size:8pt; font-weight:700; letter-spacing:.1em; text-transform:uppercase; color:var(--mata); text-align:right; }}
  .titulo {{ font-family:'Archivo'; font-weight:800; font-stretch:87%; font-size:24pt; line-height:1.04;
             letter-spacing:-.01em; color:var(--mata-escuro); margin:5mm 0 1.5mm; }}
  .titulo em {{ font-style:normal; color:var(--rio); }}
  .subtitulo {{ font-family:'Archivo'; font-size:10.5pt; font-weight:500; color:var(--cinza); margin-bottom:3mm; }}

  /* ── seções ── */
  .sec {{ margin-top:4.6mm; }}
  .sec-eyebrow {{ font-family:'Archivo'; font-size:8.5pt; font-weight:700; letter-spacing:.16em;
                  text-transform:uppercase; color:var(--mata); display:flex; align-items:center; gap:8px; margin-bottom:3.5mm; }}
  .sec-eyebrow::after {{ content:''; flex:1; height:1px; background:var(--fio); }}
  .lead {{ font-size:12.5pt; line-height:1.72; }}
  .lead b {{ font-weight:700; color:var(--mata-escuro); }}
  .lead .neg {{ color:var(--rio); font-weight:700; }}
  p + p {{ margin-top:3mm; }}
  .nota-analise {{ font-size:10.5pt; color:var(--cinza); border-left:2.5px solid var(--fio); padding-left:4mm; margin-top:4mm; }}

  /* ── o curso do recurso (hero) ── */
  .rio-wrap {{ margin:5mm 0 2mm; }}
  .rio-band {{ display:flex; height:15mm; border-radius:3px; overflow:hidden; }}
  .rio-seg {{ position:relative; display:flex; align-items:center; justify-content:center; }}
  .rio-seg .v {{ font-family:'Archivo'; font-weight:800; font-size:12.5pt; color:#fff; letter-spacing:-.01em; }}
  .rio-pago   {{ background:linear-gradient(180deg,#237A52,#1B5E3F); }}
  .rio-emp    {{ background:linear-gradient(180deg,#E0973B,#C97F22); }}
  .rio-dem    {{ background:linear-gradient(180deg,#CBD3CC,#B8C2BA); }}
  .rio-dem .v {{ color:var(--tinta); }}
  .rio-leg {{ display:flex; margin-top:2.5mm; font-family:'Archivo'; font-size:8.8pt; }}
  .rio-leg > div {{ padding-right:4mm; min-width:0; position:relative; }}
  .rio-leg > div::before {{ content:''; position:absolute; top:-2.5mm; left:0; width:6mm; height:1.5px; background:currentColor; opacity:.5; }}
  .rio-leg .l-t {{ font-weight:700; text-transform:uppercase; letter-spacing:.06em; font-size:7.5pt; }}
  .rio-leg .l-v {{ font-weight:600; color:var(--tinta); font-variant-numeric:tabular-nums; }}
  .rio-leg .l-t {{ line-height:1.25; }}
  .l-pago {{ color:var(--mata); }} .l-emp {{ color:var(--ambar); }} .l-dem {{ color:var(--cinza-claro); }}
  .l-pago .l-t {{ color:var(--mata); }} .l-emp .l-t {{ color:var(--ambar); }} .l-dem .l-t {{ color:var(--cinza-claro); }}
  .rio-marco {{ font-family:'Archivo'; font-size:8pt; color:var(--cinza); margin-top:2mm; }}
  .rio-marco b {{ color:var(--mata-escuro); }}

  /* ── KPIs ── */
  .kpis {{ display:flex; gap:3mm; margin-top:5mm; }}
  .kpi {{ flex:1; background:var(--painel); border-radius:3px; padding:3.2mm 3mm 2.8mm; text-align:center; }}
  .kpi-v {{ font-family:'Archivo'; font-weight:800; font-size:15pt; letter-spacing:-.02em; font-variant-numeric:tabular-nums; }}
  .kpi-l {{ font-family:'Archivo'; font-size:7.2pt; font-weight:600; letter-spacing:.05em; text-transform:uppercase; color:var(--cinza); margin-top:.8mm; line-height:1.3; }}

  /* ── tabelas ── */
  table {{ width:100%; border-collapse:collapse; font-family:'Archivo'; font-size:9.3pt; }}
  th {{ text-align:left; font-size:7.5pt; font-weight:700; letter-spacing:.09em; text-transform:uppercase;
       color:var(--cinza); padding:0 3mm 1.8mm 0; border-bottom:1.5px solid var(--tinta); }}
  td {{ padding:1.9mm 3mm 1.9mm 0; border-bottom:.5px solid var(--fio); vertical-align:top; }}
  tr:last-child td {{ border-bottom:none; }}
  .t-nome {{ font-weight:600; }}
  .t-num  {{ text-align:right; white-space:nowrap; font-variant-numeric:tabular-nums; }}
  th.t-num {{ text-align:right; }}
  .t-nota {{ font-family:'Source Serif 4'; font-size:9.3pt; color:var(--cinza); line-height:1.5; }}
  .t-exec {{ width:34mm; }}
  .t-pct {{ font-weight:700; font-variant-numeric:tabular-nums; display:block; margin-bottom:1mm; }}
  .mini {{ height:2.2mm; background:#E3E8E2; border-radius:2px; overflow:hidden; }}
  .mini-fill {{ height:100%; background:var(--mata); border-radius:2px; }}

  /* ── frentes (concluído / análise / pendente) ── */
  .frentes {{ display:grid; grid-template-columns:1fr 1fr; gap:3mm; margin-top:1mm; }}
  .frente {{ background:var(--painel); border-radius:3px; padding:3.2mm 4mm; }}
  .frente-t {{ font-family:'Archivo'; font-weight:700; font-size:10pt; margin-bottom:.5mm; }}
  .frente-s {{ font-family:'Archivo'; font-size:8pt; color:var(--cinza); margin-bottom:2mm; }}
  .stack {{ display:flex; height:4.5mm; border-radius:2px; overflow:hidden; margin-bottom:2mm; }}
  .st-ok {{ background:#237A52; }} .st-at {{ background:#E0973B; }} .st-cr {{ background:#C2551B; }} .st-vazio {{ background:#DDE3DC; }}
  .frente-leg {{ font-family:'Archivo'; font-size:8.4pt; color:var(--cinza); line-height:1.55; }}
  .frente-leg b {{ color:var(--tinta); font-variant-numeric:tabular-nums; }}
  .frente-obs {{ font-family:'Source Serif 4'; font-size:9.2pt; color:var(--tinta); margin-top:2.2mm;
                 padding-top:2.2mm; border-top:.5px solid var(--fio); line-height:1.55; }}

  /* ── pendências ── */
  .pend {{ border-radius:3px; padding:3.2mm 4.5mm; margin-bottom:2.5mm; }}
  .pend.rio {{ background:var(--rio-claro); border-left:3px solid var(--rio); }}
  .pend-head {{ display:flex; gap:4mm; align-items:baseline; }}
  .pend-num {{ font-family:'Archivo'; font-weight:800; font-size:17pt; color:var(--rio); min-width:9mm; font-variant-numeric:tabular-nums; }}
  .pend-t {{ font-family:'Archivo'; font-weight:700; font-size:10.3pt; }}
  .pend-s {{ font-family:'Archivo'; font-size:8.6pt; color:#8A4415; }}
  .pend-m {{ font-family:'Archivo'; font-size:8.6pt; color:var(--cinza); margin-top:1.8mm; padding-left:13mm; line-height:1.6; }}

  /* ── rodapé ── */
  /* cards de instrumento */
  .ci {{ background:var(--painel); border-radius:3px; padding:3.4mm 4mm; margin-bottom:2.6mm; }}
  .ci-top {{ display:flex; justify-content:space-between; align-items:baseline; margin-bottom:2mm; }}
  .ci-nome {{ font-family:'Archivo'; font-weight:700; font-size:10.5pt; color:var(--mata-escuro); }}
  .ci-n {{ font-family:'Archivo'; font-weight:500; font-size:8pt; color:var(--cinza); margin-left:2.5mm; }}
  .ci-prev {{ font-family:'Archivo'; font-weight:700; font-size:10.5pt; }}
  .ci-bar {{ display:flex; height:4.6mm; background:#E3E8E2; border-radius:2px; overflow:hidden; }}
  .ci-pago {{ background:#237A52; }} .ci-emp {{ background:#E0973B; }}
  .ci-leg {{ display:flex; flex-wrap:wrap; gap:4mm; align-items:baseline; font-family:'Archivo';
             font-size:8.2pt; color:var(--cinza); margin-top:1.8mm; }}
  .ci-leg b {{ color:var(--tinta); }}
  .ci-fin {{ font-family:'Source Serif 4'; font-size:8.8pt; color:var(--cinza); flex:1; min-width:60mm; }}
  .dot {{ display:inline-block; width:2mm; height:2mm; border-radius:50%; margin-right:1.2mm; }}
  .dp {{ background:#237A52; }} .de {{ background:#E0973B; }}
  .et-wrap {{ margin-top:2.4mm; padding-top:2.2mm; border-top:.5px solid var(--fio); }}
  .et-lbl {{ font-family:'Archivo'; font-size:7.4pt; font-weight:700; letter-spacing:.1em;
             text-transform:uppercase; color:var(--cinza-claro); margin-bottom:1.4mm; }}
  .et {{ display:flex; gap:2.6mm; align-items:baseline; padding:.7mm 0; }}
  .et-n {{ font-family:'Archivo'; font-weight:700; font-size:8.6pt; color:var(--mata);
           min-width:5mm; text-align:right; }}
  .et-d {{ font-family:'Archivo'; font-size:8.6pt; color:var(--tinta); }}
  .et-m {{ display:block; font-family:'Source Serif 4'; font-size:8.4pt; color:var(--cinza); line-height:1.45; }}
  /* execução regional + mapa */
  .mapa {{ margin-bottom:4mm; }}
  .reg-tab {{ width:100%; }}
  .mapa svg {{ width:100%; height:auto; display:block; }}
  .mp-leg {{ display:flex; flex-wrap:wrap; gap:2.5mm; justify-content:center; margin-top:1.5mm;
             font-family:'Archivo'; font-size:7.2pt; color:var(--cinza); }}
  .mp-item {{ display:flex; align-items:center; gap:1mm; }}
  .mp-item i {{ width:2.6mm; height:2.6mm; border-radius:1px; border:.3px solid #CBD5CD; display:inline-block; }}
  .mapa-cap {{ font-family:'Archivo'; font-size:7.4pt; color:var(--cinza-claro); margin-top:1.4mm;
               line-height:1.4; text-align:center; }}
  .at-et {{ display:block; font-family:'Archivo'; font-size:7.2pt; color:var(--cinza-claro);
            text-transform:uppercase; letter-spacing:.06em; margin-top:.4mm; }}
  .rodape {{ margin-top:10mm; padding-top:2mm; display:flex; justify-content:space-between;
             font-family:'Archivo'; font-size:7.5pt; color:var(--cinza-claro); border-top:.5px solid var(--fio); }}

  /* ── impressão ── */
  .btn-print {{ position:fixed; top:14px; right:14px; z-index:9; font-family:'Archivo'; font-weight:700; font-size:13px;
                background:var(--mata); color:#fff; border:0; border-radius:22px; padding:10px 18px; cursor:pointer;
                box-shadow:0 3px 10px rgba(15,61,42,.35); }}
  .btn-print:hover {{ background:var(--mata-escuro); }}
  @page {{ size:A4; margin:15mm 16mm 13mm; }}
  @media print {{
    body {{ background:#fff; font-size:10.9pt; }}
    .pagina {{ margin:0; padding:0; box-shadow:none; width:auto; }}
    .btn-print {{ display:none; }}
    .sec-label, .frente, .pend, .rio-wrap, .kpis, tr, thead, .et, .reg-tab tr {{
      page-break-inside:avoid; break-inside:avoid; }}
    .sec-eyebrow {{ break-after:avoid; page-break-after:avoid; }}
    thead {{ display:table-header-group; }}
    .ci {{ break-inside:auto; }}
    .ci-top, .ci-bar, .ci-leg {{ break-inside:avoid; break-after:avoid; }}
    p, .marco, .nota-analise {{ orphans:3; widows:3; }}
    .rodape {{ margin-top:8mm; }}
    .t-nota {{ font-size:8.8pt; line-height:1.4; }}
    .subtitulo {{ font-size:9.4pt; }}
    .frente-obs, .fr-l {{ line-height:1.45; }}
    td {{ padding-top:1.6mm; padding-bottom:1.6mm; }}
  }}
  @media screen and (max-width:820px) {{ .pagina {{ width:auto; min-height:auto; padding:8mm 6mm; }} .frentes {{ grid-template-columns:1fr; }} .kpis {{ flex-wrap:wrap; }} .kpi {{ min-width:28%; }} }}
</style>
</head>
<body>
<button class="btn-print" onclick="window.print()">🖨&nbsp; Imprimir / salvar PDF</button>

<!-- ═════════════════ PÁGINA 1 ═════════════════ -->
<div class="pagina">
  <div class="cabo">
    <div class="cabo-org">Secretaria de Estado de Saúde de Minas Gerais<br>Subsecretaria de Regionalização · CMIR</div>
    <div class="cabo-num">Boletim de acompanhamento<br>Edição {edicao} · {hoje.strftime("%d/%m/%Y")}</div>
  </div>
  <hr class="fio-rio">
  <div class="titulo">Plano de Ação em Saúde<br>do <em>Rio Doce</em></div>
  <div class="subtitulo arch">Acompanhamento da execução — {n_benef} beneficiários · {n_linhas} ações monitoradas · posição de {data_ext}</div>

  <div class="sec">
    <div class="sec-eyebrow arch">O curso do recurso</div>
    <div class="rio-wrap">
      <div class="rio-band">
        <div class="rio-seg rio-pago" style="width:{p_pago:.1f}%"><span class="v">{pct(p_pago)}</span></div>
        <div class="rio-seg rio-emp" style="width:{p_emp:.1f}%"><span class="v">{pct(p_emp)}</span></div>
        <div class="rio-seg rio-dem" style="flex:1"><span class="v">{pct(p_dem)}</span></div>
      </div>
      <div class="rio-leg">
        <div class="l-pago" style="width:{p_pago:.1f}%"><div class="l-t">Executado</div><div class="l-v">{brl(pago)}</div></div>
        <div class="l-emp" style="width:{p_emp:.1f}%"><div class="l-t">Empenhado · aguarda pagamento</div><div class="l-v">{brl(emp)}</div></div>
        <div class="l-dem" style="flex:1"><div class="l-t">Em etapas anteriores</div><div class="l-v">{brl(demais)}</div></div>
      </div>
      <div class="rio-marco">Do total de <b>{brl(tot, 0)}</b> do plano, <b>{brl(comprometido, 0)}</b> ({pct(p_compr)}) já estão pagos ou
      empenhados. Os pagamentos alcançam <b>{n_benef_pagos} dos {n_benef} beneficiários</b> — {len(benef_100)} deles com repasses
      integralmente quitados.</div>
    </div>
  </div>

  <div class="sec">
    <div class="sec-eyebrow arch">Execução por instrumento</div>
    {cards_instr}
  </div>

  <div class="sec">
    <div class="sec-eyebrow arch" style="color:var(--rio)">Pendências</div>
    {pendencias}
  </div>

  {f"""<div class="sec">
    <div class="sec-eyebrow arch">Atualizações da semana</div>
    <table>
      <thead><tr><th>Município</th><th>Beneficiário final</th><th>Tipo de aplicação</th><th>Status</th><th class="t-num">Data</th></tr></thead>
      <tbody>{linhas_atu}</tbody>
    </table>
    <p class="nota-analise">{len(atualizacoes)} registro{"s" if len(atualizacoes) != 1 else ""} com atualização nos últimos {JANELA_DIAS} dias.</p>
  </div>""" if atualizacoes else ""}

  <div class="sec">
    <div class="sec-eyebrow arch">Execução regional</div>
    {mapa_html}
    <table class="reg-tab">
      <thead><tr><th>Microrregião de saúde</th><th class="t-num">Previsto</th><th class="t-num">Pago</th><th>Execução</th></tr></thead>
      <tbody>{linhas_micro}</tbody>
    </table>
  </div>

  <div class="rodape">
    <span>Fonte: Base Consolidada do Plano Rio Doce</span>
    <span>Elaboração: CMIR / Subsecretaria de Regionalização — SES-MG</span>
  </div>
</div>
</body>
</html>"""

with open(OUT, "w", encoding="utf-8") as f:
    f.write(html)
print(f"gerado: {OUT} ({len(html):,} chars)")
